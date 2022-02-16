import openpyxl
from matplotlib import pyplot as plt
from cal_results import cal_BDrate


def save_results_in_fig(results, qps, perf_result_path):
    for seq_name in results:
        plt.figure()
        label = []
        for encoder in results[seq_name]:
            plt.title('Encode Time')
            plt.xlabel('QP')
            plt.ylabel('Time/s')
            plt.plot(qps, results[seq_name][encoder]['encode_time'])
            label.append(encoder)
        plt.legend(label)
        plt.savefig(perf_result_path + '/' + seq_name + '_encode_time.png')

        plt.figure()
        label = []
        for encoder in results[seq_name]:
            plt.title('Decode Time')
            plt.xlabel('QP')
            plt.ylabel('Time/s')
            plt.plot(qps, results[seq_name][encoder]['decode_time'])
            label.append(encoder)
        plt.legend(label)
        plt.savefig(perf_result_path + '/' + seq_name + '_decode_time.png')

        plt.figure()
        label = []
        for encoder in results[seq_name]:
            plt.title('Rate Distortion Performance')
            plt.xlabel('bits/kbps')
            plt.ylabel('PSNR/dB')
            plt.plot(results[seq_name][encoder]['bits'], results[seq_name][encoder]['psnr'])
            label.append(encoder)
        plt.legend(label)
        plt.savefig(perf_result_path + '/' + seq_name + '_rate_distortion_performance.png')


def save_results_in_xlsx(results, qps, perf_result_path, encoders):
    workbook = openpyxl.Workbook()
    encoders_num = len(encoders)
    qps_num = len(qps)
    current_encoder = []
    side_style = openpyxl.styles.Side(border_style='medium')
    color = '97FFFF'
    for encoder_index in range(encoders_num):
        current_encoder.clear()
        current_encoder.append(encoders[0])
        if encoders_num > 1 and encoder_index == 0:
            continue
        elif encoders_num > 1:
            current_encoder.append(encoders[encoder_index])
        if encoder_index > 1:
            sheet = workbook.create_sheet(encoders[encoder_index])
        else:
            sheet = workbook.active
        for index in range(len(current_encoder)):
            sheet.title = current_encoder[index]
            sheet.merge_cells('C1:E1')
            sheet.merge_cells('F1:H1')
            start_row = 3
            start_col = 3
            seq_index = 0
            step = 3
            sheet.cell(1, start_col + index * step).value = current_encoder[index]
            sheet.cell(1, start_col + index * step).border = openpyxl.styles.Border(left=side_style, right=side_style, top=side_style, bottom=side_style)
            sheet.cell(1, start_col + index * step + 1).border = openpyxl.styles.Border(left=side_style, right=side_style, top=side_style, bottom=side_style)
            sheet.cell(1, start_col + index * step + 2).border = openpyxl.styles.Border(left=side_style, right=side_style, top=side_style, bottom=side_style)
            sheet.cell(2, 9).value = 'BD-rate'
            sheet.cell(2, 9).border = openpyxl.styles.Border(left=side_style, right=side_style, top=side_style, bottom=side_style)
            sheet.cell(2, 10).value = 'Time'
            sheet.cell(2, 10).border = openpyxl.styles.Border(left=side_style, right=side_style, top=side_style, bottom=side_style)
            sheet.cell(2, 2).value = 'QP'
            sheet.cell(2, 2).border = openpyxl.styles.Border(left=side_style, right=side_style, top=side_style, bottom=side_style)
            sheet.cell(2, start_col + index * step).value = 'kbps'
            sheet.cell(2, start_col + index * step).border = openpyxl.styles.Border(left=side_style, right=side_style, top=side_style, bottom=side_style)
            sheet.cell(2, start_col + index * step + 1).value = 'Y PSNR'
            sheet.cell(2, start_col + index * step + 1).border = openpyxl.styles.Border(left=side_style, right=side_style, top=side_style, bottom=side_style)
            sheet.cell(2, start_col + index * step + 2).value = 'Enc T [s]'
            sheet.cell(2, start_col + index * step + 2).border = openpyxl.styles.Border(left=side_style, right=side_style, top=side_style, bottom=side_style)
            for seq_name in results:
                sheet.merge_cells('A' + str(start_row + seq_index * qps_num) + ':A' + str(start_row + (seq_index + 1) * qps_num - 1))
                sheet.cell(start_row + seq_index * qps_num, 1).value = seq_name
                for i in range(qps_num):
                    sheet.cell(start_row + seq_index * qps_num + i, 1).border = openpyxl.styles.Border(left=side_style, right=side_style, top=side_style, bottom=side_style)
                for i in range(start_row + seq_index * qps_num, start_row + (seq_index + 1) * qps_num):
                    sheet.cell(i, 2).value = qps[(i - start_row) % qps_num]
                    sheet.cell(i, start_col + index * step).value = results[seq_name][current_encoder[index]]['bits'][(i - start_row) % qps_num]
                    sheet.cell(i, start_col + index * step).border = openpyxl.styles.Border(left=side_style)
                    sheet.cell(i, start_col + index * step).fill = openpyxl.styles.PatternFill("solid", fgColor=color)
                    sheet.cell(i, start_col + index * step + 1).value = results[seq_name][current_encoder[index]]['psnr'][(i - start_row) % qps_num]
                    sheet.cell(i, start_col + index * step + 1).fill = openpyxl.styles.PatternFill("solid", fgColor=color)
                    sheet.cell(i, start_col + index * step + 2).value = results[seq_name][current_encoder[index]]['encode_time'][(i - start_row) % qps_num]
                    sheet.cell(i, start_col + index * step + 2).border = openpyxl.styles.Border(right=side_style)
                    sheet.cell(i, start_col + index * step + 2).fill = openpyxl.styles.PatternFill("solid", fgColor=color)
                sheet.cell(i, 2).border = openpyxl.styles.Border(bottom=side_style)
                sheet.cell(i, start_col + index * step).border = openpyxl.styles.Border(left=side_style, bottom=side_style)
                sheet.cell(i, start_col + index * step + 1).border = openpyxl.styles.Border(bottom=side_style)
                sheet.cell(i, start_col + index * step + 2).border = openpyxl.styles.Border(bottom=side_style)
                if index > 0:
                    sheet.merge_cells('I' + str(start_row + seq_index * qps_num) + ':I' + str(start_row + (seq_index + 1) * qps_num - 1))
                    sheet.cell(start_row + seq_index * qps_num, 9).value = results[seq_name][current_encoder[index]]['BDrate']
                    sheet.cell(start_row + seq_index * qps_num, 9).number_format = '0.00%'
                    sheet.cell(start_row + seq_index * qps_num, 9).font = openpyxl.styles.Font(bold=True)
                    if results[seq_name][current_encoder[index]]['BDrate'] < 0:
                        sheet.cell(start_row + seq_index * qps_num, 9).fill = openpyxl.styles.PatternFill("solid", fgColor='98FB98')
                    else:
                        sheet.cell(start_row + seq_index * qps_num, 9).fill = openpyxl.styles.PatternFill("solid", fgColor='EEB4B4')
                    sheet.merge_cells('J' + str(start_row + seq_index * qps_num) + ':J' + str(start_row + (seq_index + 1) * qps_num - 1))
                    sheet.cell(start_row + seq_index * qps_num, 10).value = results[seq_name][current_encoder[index]]['time_saving']
                    sheet.cell(start_row + seq_index * qps_num, 10).number_format = '0.00%'
                    sheet.cell(start_row + seq_index * qps_num, 10).font = openpyxl.styles.Font(bold=True)
                    if results[seq_name][current_encoder[index]]['time_saving'] < 0:
                        sheet.cell(start_row + seq_index * qps_num, 10).fill = openpyxl.styles.PatternFill("solid", fgColor='98FB98')
                    else:
                        sheet.cell(start_row + seq_index * qps_num, 10).fill = openpyxl.styles.PatternFill("solid", fgColor='EEB4B4')
                    for i in range(qps_num):
                        sheet.cell(start_row + seq_index * qps_num + i, 9).border = openpyxl.styles.Border(left=side_style, right=side_style, top=side_style, bottom=side_style)
                        sheet.cell(start_row + seq_index * qps_num + i, 10).border = openpyxl.styles.Border(left=side_style, right=side_style, top=side_style, bottom=side_style)
                seq_index = seq_index + 1

    for sheet in workbook:
        row_num = str(sheet.max_row)  # 总行数
        letter = openpyxl.utils.get_column_letter(sheet.max_column)  # 最后一列的字母
        cells = sheet['A1:' + letter + row_num]  # 全部单元格范围
        for i in cells:
            for j in i:
                j.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')  # 居中
                j.font = openpyxl.styles.Font(name='Times New Roman')
        sheet.column_dimensions['A'].width = 20
    workbook.save(perf_result_path + '/../results.xlsx')


def save_results(results, qps, perf_result_path, encoders):
    # save rate distortion figure and encode/decode time figure
    save_results_in_fig(results, qps, perf_result_path)
    # calculate BDrate
    encoders_num = len(encoders)
    if encoders_num >= 2:
        cal_BDrate(results)
    # save results in xlsx
    save_results_in_xlsx(results, qps, perf_result_path, encoders)
